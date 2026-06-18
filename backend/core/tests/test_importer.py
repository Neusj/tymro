"""Tests del importador de datos (F0: motor + Disciplinas).

Cubre el ciclo completo plantilla → validate → commit, el aislamiento
multitenant, los permisos por rol y la resolución de FK del motor.
"""
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from core.models import Branch, ClassType, Discipline

pytestmark = pytest.mark.django_db

PASSWORD = 'Passw0rd2026'
LOGIN_URL = '/api/login/'
ENTITIES_URL = '/api/imports/entities/'
TEMPLATE_URL = '/api/imports/disciplines/template/'
VALIDATE_URL = '/api/imports/disciplines/validate/'
COMMIT_URL = '/api/imports/disciplines/commit/'

HEADERS = ['Nombre', 'Descripción', 'Activa']


def build_xlsx_bytes(rows, headers=None, sheet_name='Datos'):
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers if headers is not None else HEADERS)
    for row in rows:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def as_upload(file_bytes, filename='disciplinas.xlsx'):
    return SimpleUploadedFile(
        filename,
        file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def build_xlsx(rows, headers=None, sheet_name='Datos', filename='disciplinas.xlsx'):
    # OJO: openpyxl guarda un timestamp interno, así que dos llamadas con el
    # mismo contenido NO producen bytes idénticos. Para el ciclo validate→commit
    # (token atado al sha256) usa build_xlsx_bytes + as_upload con LOS MISMOS bytes.
    return as_upload(build_xlsx_bytes(rows, headers, sheet_name), filename)


def login(api_client, username):
    resp = api_client.post(LOGIN_URL, {'username': username, 'password': PASSWORD}, format='json')
    assert resp.status_code == 200, resp.content
    api_client.credentials(HTTP_AUTHORIZATION=f"Token {resp.json()['token']}")


@pytest.fixture
def org_a(make_organization):
    return make_organization('Gimnasio A')


@pytest.fixture
def org_b(make_organization):
    return make_organization('Gimnasio B')


@pytest.fixture
def admin_a(make_user, org_a):
    return make_user('admin_a', organization=org_a, role='gym_admin', email='admin_a@tymro.cl')


@pytest.fixture
def admin_b(make_user, org_b):
    return make_user('admin_b', organization=org_b, role='gym_admin', email='admin_b@tymro.cl')


@pytest.fixture
def superadmin(make_user):
    return make_user('root', organization=None, role='superadmin', email='root@tymro.cl')


def validate_file(api_client, upload, extra=None):
    data = {'file': upload}
    data.update(extra or {})
    return api_client.post(VALIDATE_URL, data, format='multipart')


def commit_file(api_client, upload, token, extra=None):
    data = {'file': upload}
    if token is not None:
        data['token'] = token
    data.update(extra or {})
    return api_client.post(COMMIT_URL, data, format='multipart')


def full_import(api_client, rows, extra=None):
    """validate + commit del MISMO archivo (mismos bytes); devuelve la respuesta del commit."""
    file_bytes = build_xlsx_bytes(rows)
    resp = validate_file(api_client, as_upload(file_bytes), extra)
    assert resp.status_code == 200, resp.content
    token = resp.json()['token']
    return commit_file(api_client, as_upload(file_bytes), token, extra)


# ---------------------------------------------------------------- catálogo

def test_entities_catalog_returns_disciplines_spec(api_client, admin_a):
    login(api_client, 'admin_a')
    resp = api_client.get(ENTITIES_URL)
    assert resp.status_code == 200
    entities = resp.json()['entities']
    disciplines = next(e for e in entities if e['slug'] == 'disciplines')
    assert disciplines['label'] == 'Disciplinas'
    assert disciplines['dependencies'] == []
    assert disciplines['instructions']
    labels = {f['label']: f for f in disciplines['fields']}
    assert labels['Nombre']['required'] is True
    assert labels['Nombre']['max_length'] == 120
    assert labels['Descripción']['required'] is False
    assert labels['Activa']['choices'] == ['Sí', 'No']


@pytest.mark.parametrize('role', ['manager', 'monitor', 'teacher', 'student'])
def test_forbidden_for_non_admin_roles(api_client, make_user, org_a, role):
    make_user(f'user_{role}', organization=org_a, role=role, email=f'{role}@tymro.cl')
    login(api_client, f'user_{role}')
    assert api_client.get(ENTITIES_URL).status_code == 403
    assert api_client.get(TEMPLATE_URL).status_code == 403
    assert validate_file(api_client, build_xlsx([['Yoga', '', 'Sí']])).status_code == 403
    assert commit_file(api_client, build_xlsx([['Yoga', '', 'Sí']]), 'x').status_code == 403


def test_anonymous_gets_401(api_client):
    assert api_client.get(ENTITIES_URL).status_code == 401


def test_unknown_entity_returns_404(api_client, admin_a):
    login(api_client, 'admin_a')
    assert api_client.get('/api/imports/no-existe/template/').status_code == 404


# ---------------------------------------------------------------- plantilla

def test_template_download_three_sheets(api_client, admin_a):
    from openpyxl import load_workbook

    login(api_client, 'admin_a')
    resp = api_client.get(TEMPLATE_URL)
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp['Content-Type']
    assert 'plantilla_disciplines.xlsx' in resp['Content-Disposition']

    workbook = load_workbook(BytesIO(resp.content))
    assert set(workbook.sheetnames) == {'Datos', 'Instrucciones', 'Referencias'}
    data_sheet = workbook['Datos']
    headers = [cell.value for cell in data_sheet[1]]
    assert headers == HEADERS
    assert all(cell.font.bold for cell in data_sheet[1])
    assert data_sheet['A2'].value == 'Yoga'  # fila de ejemplo


# ---------------------------------------------------------------- validate

def test_validate_ok_returns_token_and_persists_nothing(api_client, admin_a):
    login(api_client, 'admin_a')
    upload = build_xlsx([['Yoga', 'Suave', 'Sí'], ['Crossfit', '', 'No'], ['Funcional', '', '']])
    resp = validate_file(api_client, upload)
    assert resp.status_code == 200
    body = resp.json()
    assert body['can_commit'] is True
    assert body['token']
    assert body['summary'] == {
        'total_rows': 3, 'valid': 3, 'duplicates_in_file': 0,
        'duplicates_in_db': 0, 'errors': 0, 'will_create': 3,
        'updated': 0, 'unchanged': 0,
    }
    assert Discipline.objects.count() == 0


def test_validate_reports_spanish_errors_with_row_and_column(api_client, admin_a):
    login(api_client, 'admin_a')
    upload = build_xlsx([
        ['', 'sin nombre', 'Sí'],
        ['x' * 121, '', 'Sí'],
        ['Pilates', '', 'quizás'],
    ])
    resp = validate_file(api_client, upload)
    body = resp.json()
    assert resp.status_code == 200
    assert body['can_commit'] is False
    assert body['summary']['errors'] == 3

    by_row = {r['row']: r for r in body['rows']}
    error_2 = by_row[2]['errors'][0]
    assert error_2['column'] == 'Nombre'
    assert error_2['message'] == "El campo 'Nombre' es obligatorio."
    error_3 = by_row[3]['errors'][0]
    assert 'supera el máximo de 120 caracteres' in error_3['message']
    error_4 = by_row[4]['errors'][0]
    assert error_4['column'] == 'Activa'
    assert "Usa 'Sí' o 'No'" in error_4['message']


def test_validate_missing_column_is_file_error(api_client, admin_a):
    # Falta una columna OBLIGATORIA ('Nombre') -> error de archivo. Las columnas
    # opcionales ausentes sí se toleran (ver test_validate_optional_column_can_be_absent).
    login(api_client, 'admin_a')
    upload = build_xlsx([['Una descripción', 'Sí']], headers=['Descripción', 'Activa'])
    resp = validate_file(api_client, upload)
    assert resp.status_code == 400
    assert "Falta la columna 'Nombre'" in resp.json()['detail']


def test_validate_optional_column_can_be_absent(api_client, admin_a):
    # 'Descripción' es opcional: un archivo sin esa columna se valida igual.
    login(api_client, 'admin_a')
    upload = build_xlsx([['Yoga', 'Sí']], headers=['Nombre', 'Activa'])
    resp = validate_file(api_client, upload)
    assert resp.status_code == 200, resp.content
    assert resp.json()['can_commit'] is True


def test_validate_rejects_non_xlsx_and_garbage(api_client, admin_a):
    login(api_client, 'admin_a')
    csv_file = SimpleUploadedFile('datos.csv', b'Nombre\nYoga', content_type='text/csv')
    resp = validate_file(api_client, csv_file)
    assert resp.status_code == 400
    assert '.xlsx' in resp.json()['detail']

    garbage = SimpleUploadedFile('datos.xlsx', b'esto no es un excel', content_type='text/plain')
    resp = validate_file(api_client, garbage)
    assert resp.status_code == 400
    assert 'no es un Excel válido' in resp.json()['detail']


def test_validate_rejects_oversized_file(api_client, admin_a, monkeypatch):
    monkeypatch.setattr('core.importer.engine.MAX_FILE_SIZE', 100)
    login(api_client, 'admin_a')
    resp = validate_file(api_client, build_xlsx([['Yoga', '', 'Sí']]))
    assert resp.status_code == 400
    assert 'tamaño máximo' in resp.json()['detail']


def test_validate_rejects_too_many_rows(api_client, admin_a):
    login(api_client, 'admin_a')
    rows = [[f'Disciplina {i}', '', 'Sí'] for i in range(1001)]
    resp = validate_file(api_client, build_xlsx(rows))
    assert resp.status_code == 400
    assert 'más de 1000 filas' in resp.json()['detail']


def test_validate_missing_data_sheet(api_client, admin_a):
    login(api_client, 'admin_a')
    upload = build_xlsx([['Yoga', '', 'Sí']], sheet_name='Hoja1')
    resp = validate_file(api_client, upload)
    assert resp.status_code == 400
    assert 'hoja "Datos"' in resp.json()['detail']


def test_validate_dedup_within_file_case_insensitive(api_client, admin_a):
    login(api_client, 'admin_a')
    resp = validate_file(api_client, build_xlsx([['Yoga', '', 'Sí'], [' yoga ', '', 'No']]))
    body = resp.json()
    assert body['summary']['will_create'] == 1
    assert body['summary']['duplicates_in_file'] == 1
    dup = next(r for r in body['rows'] if r['status'] == 'duplicado_archivo')
    assert 'fila 2' in dup['note']
    assert body['can_commit'] is True  # duplicado no es error


def test_validate_dedup_against_db_case_insensitive_trim(api_client, admin_a, org_a):
    Discipline.objects.create(organization=org_a, name='Yoga')
    login(api_client, 'admin_a')
    resp = validate_file(api_client, build_xlsx([[' YOGA ', '', 'Sí'], ['Pilates', '', 'Sí']]))
    body = resp.json()
    assert body['summary']['duplicates_in_db'] == 1
    assert body['summary']['will_create'] == 1
    dup = next(r for r in body['rows'] if r['status'] == 'duplicado_existente')
    assert 'se omitirá' in dup['note']


# ---------------------------------------------------------------- commit

def test_commit_creates_rows_and_is_idempotent(api_client, admin_a, org_a):
    login(api_client, 'admin_a')
    rows = [['Yoga', 'Suave', 'Sí'], ['Crossfit', '', 'No']]
    resp = full_import(api_client, rows)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 2
    assert Discipline.objects.filter(organization=org_a).count() == 2
    crossfit = Discipline.objects.get(organization=org_a, name='Crossfit')
    assert crossfit.is_active is False

    # Re-importar el mismo archivo: no duplica nada (on_conflict=skip)
    resp = full_import(api_client, rows)
    assert resp.status_code == 201
    assert resp.json()['created'] == 0
    assert resp.json()['skipped_duplicates'] == 2
    assert Discipline.objects.filter(organization=org_a).count() == 2


def test_commit_atomic_rollback_on_any_error(api_client, admin_a):
    login(api_client, 'admin_a')
    rows = [['Yoga', '', 'Sí'], ['Pilates', '', 'Sí'], ['Boxeo', '', 'Sí'],
            ['Spinning', '', 'Sí'], ['', 'sin nombre', 'Sí']]
    file_bytes = build_xlsx_bytes(rows)
    token = validate_file(api_client, as_upload(file_bytes)).json()['token']
    resp = commit_file(api_client, as_upload(file_bytes), token)
    assert resp.status_code == 400
    body = resp.json()
    assert 'No se importó ningún dato' in body['detail']
    assert all(r['status'] == 'error' for r in body['rows'])
    assert Discipline.objects.count() == 0  # rollback total


def test_commit_requires_matching_token(api_client, admin_a):
    login(api_client, 'admin_a')
    token = validate_file(api_client, build_xlsx([['Yoga', '', 'Sí']])).json()['token']

    # Otro archivo con el mismo token → 400
    resp = commit_file(api_client, build_xlsx([['Pilates', '', 'Sí']]), token)
    assert resp.status_code == 400
    assert 'no coincide' in resp.json()['detail']

    # Sin token → 400
    resp = commit_file(api_client, build_xlsx([['Yoga', '', 'Sí']]), None)
    assert resp.status_code == 400
    assert 'Falta el token' in resp.json()['detail']
    assert Discipline.objects.count() == 0


def test_commit_token_expired(api_client, admin_a, monkeypatch):
    login(api_client, 'admin_a')
    file_bytes = build_xlsx_bytes([['Yoga', '', 'Sí']])
    token = validate_file(api_client, as_upload(file_bytes)).json()['token']
    monkeypatch.setattr('core.importer.engine.IMPORT_TOKEN_MAX_AGE', -1)
    resp = commit_file(api_client, as_upload(file_bytes), token)
    assert resp.status_code == 400
    assert 'expiró' in resp.json()['detail']


def test_commit_token_bound_to_organization(api_client, admin_a, admin_b):
    login(api_client, 'admin_a')
    file_bytes = build_xlsx_bytes([['Yoga', '', 'Sí']])
    token = validate_file(api_client, as_upload(file_bytes)).json()['token']

    api_client.credentials()
    login(api_client, 'admin_b')
    resp = commit_file(api_client, as_upload(file_bytes), token)
    assert resp.status_code == 400
    assert 'no coincide' in resp.json()['detail']
    assert Discipline.objects.count() == 0


# ---------------------------------------------------------------- multitenancy

def test_org_always_from_user_and_dedup_is_org_scoped(api_client, admin_a, org_a, org_b):
    # 'Yoga' existe en la org B: NO debe bloquear la importación en la org A
    Discipline.objects.create(organization=org_b, name='Yoga')
    login(api_client, 'admin_a')
    resp = full_import(api_client, [['Yoga', '', 'Sí']])
    assert resp.status_code == 201
    assert resp.json()['created'] == 1
    assert Discipline.objects.filter(organization=org_a, name='Yoga').exists()
    assert Discipline.objects.filter(organization=org_b).count() == 1


def test_gym_admin_cannot_target_other_organization(api_client, admin_a, org_b):
    login(api_client, 'admin_a')
    resp = validate_file(api_client, build_xlsx([['Yoga', '', 'Sí']]),
                         extra={'organization': org_b.id})
    assert resp.status_code == 400
    assert 'otra organización' in str(resp.json())
    assert Discipline.objects.count() == 0


def test_superadmin_requires_explicit_organization(api_client, superadmin, org_a):
    login(api_client, 'root')
    resp = validate_file(api_client, build_xlsx([['Yoga', '', 'Sí']]))
    assert resp.status_code == 400
    assert 'organización' in str(resp.json())

    resp = full_import(api_client, [['Yoga', '', 'Sí']], extra={'organization': org_a.id})
    assert resp.status_code == 201
    assert Discipline.objects.get(name='Yoga').organization_id == org_a.id


def test_superadmin_with_unknown_organization(api_client, superadmin):
    login(api_client, 'root')
    resp = validate_file(api_client, build_xlsx([['Yoga', '', 'Sí']]),
                         extra={'organization': 99999})
    assert resp.status_code == 400
    assert 'no existe' in str(resp.json())


# ---------------------------------------------------------- endurecimientos del motor

def test_validate_rejects_too_many_physical_rows_even_if_empty(api_client, admin_a):
    # XLSX con miles de filas físicas vacías: no debe iterarse completo (DoS).
    login(api_client, 'admin_a')
    rows = [['', '', '']] * 10001
    resp = validate_file(api_client, build_xlsx(rows))
    assert resp.status_code == 400
    assert 'demasiadas filas' in resp.json()['detail']


def test_coerce_email_kind_validates_format():
    from core.importer.engine import _coerce
    from core.importer.spec import FieldSpec

    field = FieldSpec(attr='email', label='Email', kind='email', required=True)
    value, error = _coerce(field, ' alumno@tymro.cl ')
    assert value == 'alumno@tymro.cl'
    assert error is None

    value, error = _coerce(field, 'no-es-un-correo')
    assert value is None
    assert 'no es válido' in error


def test_registry_rejects_fk_spec_without_org_scoping():
    from core.importer.registry import register
    from core.importer.spec import EntityImportSpec, FieldSpec, FKSpec

    spec = EntityImportSpec(
        slug='entidad-insegura', label='X', description='', model='core.Discipline',
        fields=(FieldSpec(attr='branch', label='Sucursal', kind='fk',
                          fk=FKSpec(model='core.Branch', org_field='')),),
        natural_key=('branch',),
    )
    with pytest.raises(ValueError, match='org_field'):
        register(spec)


def test_template_references_neutralize_formula_injection(org_a):
    from core.importer.spec import EntityImportSpec, FieldSpec, FKSpec
    from core.importer.templates import build_template

    ClassType.objects.create(organization=org_a, name='=HYPERLINK("http://evil";"x")')
    spec = EntityImportSpec(
        slug='tmp-formula', label='Tmp', description='', model='core.Discipline',
        fields=(FieldSpec(attr='class_type', label='Tipo de clase', kind='fk',
                          fk=FKSpec(model='core.ClassType', reference_label='tipo de clase')),),
        natural_key=('class_type',),
    )
    workbook = build_template(spec, org_a)
    cell = workbook['Referencias'].cell(row=3, column=1)
    assert cell.value == '=HYPERLINK("http://evil";"x")'
    assert cell.data_type == 's'  # texto plano, nunca fórmula


def test_commit_runs_model_full_clean(org_a):
    # Si el modelo rechaza una fila que el spec dejó pasar (aquí: spec sin
    # max_length frente a CharField(120)), el commit hace rollback total y
    # reporta el error sobre la fila/columna.
    from core.importer import engine
    from core.importer.spec import EntityImportSpec, FieldSpec

    spec = EntityImportSpec(
        slug='tmp-clean', label='Disciplinas', description='', model='core.Discipline',
        fields=(FieldSpec(attr='name', label='Nombre', kind='string', required=True),),
        natural_key=('name',),
    )
    upload = build_xlsx([['x' * 121]], headers=['Nombre'])
    file_bytes = upload.read()
    upload.seek(0)
    token = engine.issue_token(spec, org_a, file_bytes)

    with pytest.raises(engine.ImportCommitError) as exc:
        engine.run_commit(spec, org_a, upload, token)
    report = exc.value.report
    assert report.error_count == 1
    assert report.rows[0].errors[0].column == 'Nombre'
    assert Discipline.objects.count() == 0  # rollback


def test_commit_revalidates_and_skips_rows_created_meanwhile(api_client, admin_a):
    # Carrera benigna: alguien crea 'Pilates' entre validate y commit. La
    # re-validación del commit lo detecta como existente y lo omite.
    login(api_client, 'admin_a')
    file_bytes = build_xlsx_bytes([['Pilates', '', 'Sí']])
    token = validate_file(api_client, as_upload(file_bytes)).json()['token']
    Discipline.objects.create(organization=admin_a.organization, name='Pilates')
    resp = commit_file(api_client, as_upload(file_bytes), token)
    assert resp.status_code == 201
    assert resp.json()['created'] == 0
    assert Discipline.objects.filter(name='Pilates').count() == 1


# ---------------------------------------------------------------- F1: Sucursales

BRANCH_HEADERS = ['Nombre', 'Código', 'Dirección', 'Activa']


def import_entity(api_client, entity, rows, headers, extra=None):
    """Ciclo validate+commit de cualquier entidad; devuelve la respuesta del commit."""
    file_bytes = build_xlsx_bytes(rows, headers=headers)
    resp = api_client.post(f'/api/imports/{entity}/validate/',
                           {'file': as_upload(file_bytes), **(extra or {})}, format='multipart')
    assert resp.status_code == 200, resp.content
    token = resp.json()['token']
    return api_client.post(
        f'/api/imports/{entity}/commit/',
        {'file': as_upload(file_bytes), 'token': token, **(extra or {})},
        format='multipart',
    )


def test_catalog_lists_entities_in_load_order(api_client, admin_a):
    login(api_client, 'admin_a')
    entities = api_client.get(ENTITIES_URL).json()['entities']
    assert [e['slug'] for e in entities] == [
        'branches', 'disciplines', 'class-types', 'plans', 'students', 'teachers',
        'memberships', 'class-templates',
    ]
    memberships = next(e for e in entities if e['slug'] == 'memberships')
    assert memberships['dependencies'] == ['students', 'plans']


def test_branches_full_cycle_and_idempotent(api_client, admin_a, org_a, org_b):
    # 'Sede Centro' existe en la org B: no debe interferir con la org A.
    Branch.objects.create(organization=org_b, name='Sede Centro')
    login(api_client, 'admin_a')

    resp = api_client.get('/api/imports/branches/template/')
    assert resp.status_code == 200
    assert 'plantilla_branches.xlsx' in resp['Content-Disposition']

    rows = [
        ['Sede Centro', 'CEN', 'Av. Libertador 1234', 'Sí'],
        ['Sede Norte', '', '', ''],
    ]
    resp = import_entity(api_client, 'branches', rows, BRANCH_HEADERS)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 2

    centro = Branch.objects.get(organization=org_a, name='Sede Centro')
    assert centro.code == 'CEN'
    assert centro.address == 'Av. Libertador 1234'
    norte = Branch.objects.get(organization=org_a, name='Sede Norte')
    assert norte.code == ''
    assert norte.is_active is True

    resp = import_entity(api_client, 'branches', rows, BRANCH_HEADERS)
    assert resp.json()['created'] == 0
    assert resp.json()['skipped_duplicates'] == 2
    assert Branch.objects.filter(organization=org_a).count() == 2
    assert Branch.objects.filter(organization=org_b).count() == 1


def test_branches_dedup_against_db(api_client, admin_a, org_a):
    Branch.objects.create(organization=org_a, name='Sede Centro')
    login(api_client, 'admin_a')
    upload = build_xlsx([[' sede centro ', '', '', '']], headers=BRANCH_HEADERS)
    resp = api_client.post('/api/imports/branches/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['summary']['duplicates_in_db'] == 1
    assert body['summary']['will_create'] == 0


# ---------------------------------------------------------------- F1: Tipos de clase

CLASS_TYPE_HEADERS = ['Nombre', 'Descripción', 'Duración (minutos)', 'Clase privada', 'Activa']


def test_class_types_full_cycle_with_defaults(api_client, admin_a, org_a, org_b):
    # 'Clase grupal' existe en la org B: no debe interferir con la org A.
    ClassType.objects.create(organization=org_b, name='Clase grupal')
    login(api_client, 'admin_a')
    rows = [
        ['Clase grupal', 'Hasta 20 personas', 90, 'No', 'Sí'],
        ['Personalizada', '', '', 'Sí', ''],
    ]
    resp = import_entity(api_client, 'class-types', rows, CLASS_TYPE_HEADERS)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 2

    grupal = ClassType.objects.get(organization=org_a, name='Clase grupal')
    assert grupal.duration_minutes == 90
    assert grupal.is_private is False
    assert ClassType.objects.filter(organization=org_b).count() == 1  # B intacta
    personalizada = ClassType.objects.get(organization=org_a, name='Personalizada')
    assert personalizada.duration_minutes == 60  # default del spec
    assert personalizada.is_private is True
    assert personalizada.is_active is True
    assert personalizada.color == '#f97316'  # default del modelo (no se importa)


def test_class_types_duration_must_be_positive(api_client, admin_a):
    login(api_client, 'admin_a')
    upload = build_xlsx(
        [['Grupal', '', 0, '', ''], ['Express', '', 'media hora', '', ''], ['Maratón', '', 1441, '', '']],
        headers=CLASS_TYPE_HEADERS,
    )
    resp = api_client.post('/api/imports/class-types/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    by_row = {r['row']: r for r in body['rows']}
    error_2 = by_row[2]['errors'][0]
    assert error_2['column'] == 'Duración (minutos)'
    assert 'entre 1 y 1440' in error_2['message']
    error_3 = by_row[3]['errors'][0]
    assert 'no es un número válido' in error_3['message']
    assert 'entre 1 y 1440' in by_row[4]['errors'][0]['message']
    assert ClassType.objects.count() == 0


# ---------------------------------------------------------------- F2: Planes

PLAN_HEADERS = ['Nombre', 'Tipo de plan', 'Cantidad de clases', 'Clases ilimitadas',
                'Duración (días)', 'Precio', 'Visible para alumnos', 'Activo']


def test_plans_full_cycle_with_unlimited(api_client, admin_a, org_a):
    from core.models import Plan

    login(api_client, 'admin_a')
    rows = [
        ['Plan 8 clases', 'Mensual', 8, 'No', 30, 45000, 'Sí', 'Sí'],
        ['Plan libre', 'Mensual', '', 'Sí', 30, 60000, '', ''],
        ['Clase suelta', 'Clase suelta', 1, '', 7, 9000, 'Sí', 'Sí'],
    ]
    resp = import_entity(api_client, 'plans', rows, PLAN_HEADERS)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 3

    pack = Plan.objects.get(organization=org_a, name='Plan 8 clases')
    assert pack.plan_type == 'monthly'
    assert pack.total_classes == 8
    assert pack.unlimited_classes is False
    assert pack.price == 45000

    libre = Plan.objects.get(organization=org_a, name='Plan libre')
    assert libre.unlimited_classes is True
    assert libre.total_classes == 0  # forzado por derive
    assert libre.is_public is True  # default

    suelta = Plan.objects.get(organization=org_a, name='Clase suelta')
    assert suelta.plan_type == 'single_class'


def test_plans_natural_key_is_name_plus_type(api_client, admin_a, org_a):
    from core.models import Plan

    # Mismo nombre ya existente pero de OTRO tipo: no es duplicado.
    Plan.objects.create(organization=org_a, name='Promo', plan_type='pack',
                        total_classes=10, duration_days=60, price=30000)
    login(api_client, 'admin_a')
    rows = [
        ['Promo', 'Mensual', 8, 'No', 30, 45000, '', ''],   # mismo nombre, tipo distinto → crea
        ['Promo', 'Pack', 10, 'No', 60, 30000, '', ''],     # mismo nombre y tipo → omite
        ['Promo', 'Mensual', 12, 'No', 30, 50000, '', ''],  # repetido dentro del archivo → omite
    ]
    upload = build_xlsx(rows, headers=PLAN_HEADERS)
    resp = api_client.post('/api/imports/plans/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['summary']['will_create'] == 1
    assert body['summary']['duplicates_in_db'] == 1
    assert body['summary']['duplicates_in_file'] == 1
    # El preview muestra la etiqueta española del tipo, no el valor técnico.
    assert body['rows'][0]['values']['Tipo de plan'] == 'Mensual'

    resp = import_entity(api_client, 'plans', rows, PLAN_HEADERS)
    assert resp.status_code == 201
    assert resp.json()['created'] == 1
    assert Plan.objects.filter(organization=org_a, name='Promo').count() == 2


def test_plans_financial_validations(api_client, admin_a):
    from core.models import Plan

    login(api_client, 'admin_a')
    rows = [
        ['Sin clases', 'Mensual', '', 'No', 30, 45000, '', ''],      # falta cantidad
        ['Tipo malo', 'Semanal', 8, 'No', 30, 45000, '', ''],        # choice inválido
        ['Precio malo', 'Mensual', 8, 'No', 30, -100, '', ''],       # precio negativo
        ['Duración mala', 'Mensual', 8, 'No', 0, 45000, '', ''],     # duración 0
    ]
    upload = build_xlsx(rows, headers=PLAN_HEADERS)
    resp = api_client.post('/api/imports/plans/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    assert body['summary']['errors'] == 4
    by_row = {r['row']: r for r in body['rows']}
    assert "marca 'Clases ilimitadas'" in by_row[2]['errors'][0]['message']
    assert 'Las opciones son' in by_row[3]['errors'][0]['message']
    assert 'no puede ser negativo' in by_row[4]['errors'][0]['message']
    assert 'entre 1 y 3660' in by_row[5]['errors'][0]['message']
    assert Plan.objects.count() == 0


def test_plans_unlimited_overrides_filled_total_and_preview_shows_it(api_client, admin_a, org_a):
    from core.models import Plan

    login(api_client, 'admin_a')
    # Ilimitado con cantidad rellenada: derive la fuerza a 0 y el preview lo refleja.
    rows = [['Plan libre', 'Mensual', 8, 'Sí', 30, 60000, '', '']]
    file_bytes = build_xlsx_bytes(rows, headers=PLAN_HEADERS)
    resp = api_client.post('/api/imports/plans/validate/', {'file': as_upload(file_bytes)},
                           format='multipart')
    body = resp.json()
    assert body['rows'][0]['values']['Cantidad de clases'] == '0'

    resp = api_client.post('/api/imports/plans/commit/',
                           {'file': as_upload(file_bytes), 'token': body['token']},
                           format='multipart')
    assert resp.status_code == 201
    libre = Plan.objects.get(organization=org_a, name='Plan libre')
    assert libre.unlimited_classes is True
    assert libre.total_classes == 0


def test_plans_duration_upper_bound(api_client, admin_a):
    login(api_client, 'admin_a')
    upload = build_xlsx([['Eterno', 'Mensual', 8, 'No', 3661, 45000, '', '']], headers=PLAN_HEADERS)
    resp = api_client.post('/api/imports/plans/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    assert 'entre 1 y 3660' in body['rows'][0]['errors'][0]['message']


def test_plans_template_has_plan_type_dropdown(api_client, admin_a):
    from openpyxl import load_workbook

    login(api_client, 'admin_a')
    resp = api_client.get('/api/imports/plans/template/')
    assert resp.status_code == 200
    workbook = load_workbook(BytesIO(resp.content))
    references = workbook['Referencias']
    columns = {references.cell(row=2, column=col).value for col in range(1, 6)}
    assert 'Tipo de plan' in columns
    all_values = {
        references.cell(row=row, column=col).value
        for row in range(3, 10) for col in range(1, 8)
    }
    for label in ('Mensual', 'Pack', 'Clase suelta', 'Trial', 'Giftcard'):
        assert label in all_values


# ---------------------------------------------------------------- F3: Alumnos y Profesores

USER_HEADERS = ['Email', 'Nombre', 'Apellido', 'Teléfono', 'Sucursal']


def test_students_full_cycle_with_branch(api_client, admin_a, org_a):
    from django.contrib.auth import get_user_model

    Branch.objects.create(organization=org_a, name='Sede Centro')
    login(api_client, 'admin_a')
    rows = [
        ['maria.perez@gmail.com', 'María', 'Pérez', '+56 9 1234 5678', ' sede centro '],
        ['juan.soto@gmail.com', 'Juan', '', '', ''],
    ]
    resp = import_entity(api_client, 'students', rows, USER_HEADERS)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 2

    User = get_user_model()
    maria = User.objects.get(email='maria.perez@gmail.com')
    assert maria.username == 'maria.perez@gmail.com'
    assert maria.role == 'student'
    assert maria.organization_id == org_a.id
    assert maria.branch.name == 'Sede Centro'
    assert maria.has_usable_password() is False
    juan = User.objects.get(email='juan.soto@gmail.com')
    assert juan.branch is None

    # Idempotente: re-importar no duplica ni toca lo existente.
    resp = import_entity(api_client, 'students', rows, USER_HEADERS)
    assert resp.json()['created'] == 0
    assert resp.json()['skipped_duplicates'] == 2


def test_teachers_get_teacher_role(api_client, admin_a, org_a):
    from django.contrib.auth import get_user_model

    login(api_client, 'admin_a')
    resp = import_entity(api_client, 'teachers', [['coach@gym.cl', 'Pedro', 'Coach', '', '']],
                         USER_HEADERS)
    assert resp.status_code == 201
    coach = get_user_model().objects.get(email='coach@gym.cl')
    assert coach.role == 'teacher'
    assert coach.organization_id == org_a.id
    assert coach.has_usable_password() is False


def test_users_branch_must_belong_to_org(api_client, admin_a, org_b):
    Branch.objects.create(organization=org_b, name='Sede Ajena')
    login(api_client, 'admin_a')
    upload = build_xlsx([['x@y.cl', 'Ana', '', '', 'Sede Ajena']], headers=USER_HEADERS)
    resp = api_client.post('/api/imports/students/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    error = body['rows'][0]['errors'][0]
    assert error['column'] == 'Sucursal'
    assert 'No se encontró' in error['message']


def test_users_email_in_other_org_is_explicit_conflict(api_client, admin_a, admin_b):
    # admin_b@tymro.cl existe en la org B: importarlo en la org A es error
    # explícito de fila, NO un skip silencioso ni un toque a esa cuenta.
    login(api_client, 'admin_a')
    upload = build_xlsx([['ADMIN_B@tymro.cl', 'Otro', '', '', '']], headers=USER_HEADERS)
    resp = api_client.post('/api/imports/students/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    error = body['rows'][0]['errors'][0]
    assert error['column'] == 'Email'
    assert 'otra organización' in error['message']


def test_users_email_in_same_org_is_skipped_not_modified(api_client, admin_a, org_a):
    from django.contrib.auth import get_user_model

    login(api_client, 'admin_a')
    # admin_a@tymro.cl ya existe en la org A (rol gym_admin): se omite y NO se
    # degrada su rol a student.
    rows = [['admin_a@tymro.cl', 'Hacker', '', '', ''], ['nuevo@gym.cl', 'Nuevo', '', '', '']]
    resp = import_entity(api_client, 'students', rows, USER_HEADERS)
    assert resp.status_code == 201
    assert resp.json()['created'] == 1
    assert resp.json()['skipped_duplicates'] == 1
    admin = get_user_model().objects.get(email='admin_a@tymro.cl')
    assert admin.role == 'gym_admin'
    assert admin.first_name != 'Hacker'


def test_users_invalid_email_and_file_dedup(api_client, admin_a):
    login(api_client, 'admin_a')
    rows = [
        ['no-es-correo', 'Ana', '', '', ''],
        ['ana@gym.cl', 'Ana', '', '', ''],
        [' ANA@GYM.CL ', 'Ana B', '', '', ''],
    ]
    upload = build_xlsx(rows, headers=USER_HEADERS)
    resp = api_client.post('/api/imports/students/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['summary']['errors'] == 1
    assert 'no es válido' in body['rows'][0]['errors'][0]['message']
    assert body['summary']['duplicates_in_file'] == 1
    assert body['summary']['will_create'] == 1


def test_users_email_of_platform_account_is_conflict(api_client, admin_a, superadmin):
    # El superadmin no tiene organización: igual debe detectarse el conflicto
    # en validate (no recién en el commit).
    login(api_client, 'admin_a')
    upload = build_xlsx([['root@tymro.cl', 'Root', '', '', '']], headers=USER_HEADERS)
    resp = api_client.post('/api/imports/students/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    assert 'otra organización' in body['rows'][0]['errors'][0]['message']


def test_users_email_invalid_as_username_is_row_error(api_client, admin_a):
    # username=email: caracteres válidos en email pero no en username deben
    # rechazarse en validate con mensaje claro, no en commit con error técnico.
    login(api_client, 'admin_a')
    upload = build_xlsx([["o'brien@gym.cl", 'Ana', '', '', '']], headers=USER_HEADERS)
    resp = api_client.post('/api/imports/students/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    assert 'caracteres no permitidos' in body['rows'][0]['errors'][0]['message']


def test_users_template_references_only_own_branches(api_client, admin_a, org_a, org_b):
    from openpyxl import load_workbook

    Branch.objects.create(organization=org_a, name='Sede Propia')
    Branch.objects.create(organization=org_b, name='Sede Ajena')
    login(api_client, 'admin_a')
    resp = api_client.get('/api/imports/students/template/')
    assert resp.status_code == 200
    references = load_workbook(BytesIO(resp.content))['Referencias']
    all_values = {
        references.cell(row=row, column=col).value
        for row in range(1, 12) for col in range(1, 6)
    }
    assert 'Sede Propia' in all_values
    assert 'Sede Ajena' not in all_values  # aislamiento cross-org en la plantilla


# ---------------------------------------------------------------- F4: Membresías

MEMBERSHIP_HEADERS = ['Email del alumno', 'Tipo de plan', 'Nombre del plan',
                      'Fecha de inicio', 'Fecha de término', 'Clases restantes']


@pytest.fixture
def membership_setup(make_user, org_a):
    from core.models import Plan

    student = make_user('maria', organization=org_a, role='student', email='maria@gym.cl')
    plan = Plan.objects.create(
        organization=org_a, name='Plan 8 clases', plan_type='monthly',
        total_classes=8, duration_days=30, price=45000,
    )
    unlimited = Plan.objects.create(
        organization=org_a, name='Plan libre', plan_type='monthly',
        total_classes=0, unlimited_classes=True, duration_days=30, price=60000,
    )
    return {'student': student, 'plan': plan, 'unlimited': unlimited}


def test_memberships_full_cycle_carries_remaining_classes(api_client, admin_a, org_a, membership_setup):
    import datetime

    from core.models import StudentPlan

    login(api_client, 'admin_a')
    rows = [['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5]]
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_HEADERS)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 1

    membership = StudentPlan.objects.get(user=membership_setup['student'])
    assert membership.plan_id == membership_setup['plan'].id
    assert membership.total_classes == 8          # derivado del plan
    assert membership.classes_used == 3           # 8 totales - 5 restantes
    assert membership.unlimited_classes is False
    assert membership.is_active is True
    # end_date = inicio + (duración - 1), espejo del flujo Asignar plan
    assert membership.end_date == datetime.date(2026, 6, 30)
    assert membership.final_price == 45000.0

    # Idempotente: el alumno ya tiene membresía activa → se omite, no se toca.
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_HEADERS)
    assert resp.json()['created'] == 0
    assert resp.json()['skipped_duplicates'] == 1
    assert StudentPlan.objects.filter(user=membership_setup['student']).count() == 1


def test_memberships_unlimited_plan_ignores_remaining(api_client, admin_a, membership_setup):
    from core.models import StudentPlan

    login(api_client, 'admin_a')
    rows = [['maria@gym.cl', '', 'Plan libre', '2026-06-01', '', '']]
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_HEADERS)
    assert resp.status_code == 201, resp.content
    membership = StudentPlan.objects.get(user=membership_setup['student'])
    assert membership.unlimited_classes is True
    assert membership.total_classes == 0
    assert membership.classes_used == 0


def test_memberships_remaining_validations(api_client, admin_a, membership_setup):
    login(api_client, 'admin_a')
    rows = [
        ['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', ''],    # falta saldo
        ['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 9],     # saldo > total
        ['maria@gym.cl', '', 'Plan 8 clases', '2026-06-10', '2026-06-01', 5],  # término < inicio
    ]
    upload = build_xlsx(rows, headers=MEMBERSHIP_HEADERS)
    resp = api_client.post('/api/imports/memberships/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    by_row = {r['row']: r for r in body['rows']}
    assert "Indica las 'Clases restantes'" in by_row[2]['errors'][0]['message']
    assert 'entre 0 y 8' in by_row[3]['errors'][0]['message']
    assert 'no puede ser anterior' in by_row[4]['errors'][0]['message']


def test_memberships_plan_disambiguation_by_type(api_client, admin_a, org_a, membership_setup):
    from core.models import Plan, StudentPlan

    # Segundo plan 'Plan 8 clases' de otro tipo: el nombre solo es ambiguo.
    Plan.objects.create(organization=org_a, name='Plan 8 clases', plan_type='pack',
                        total_classes=8, duration_days=60, price=50000)
    login(api_client, 'admin_a')

    upload = build_xlsx([['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5]],
                        headers=MEMBERSHIP_HEADERS)
    resp = api_client.post('/api/imports/memberships/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    error = body['rows'][0]['errors'][0]
    assert error['column'] == 'Nombre del plan'
    assert "Completa la columna 'Tipo de plan'" in error['message']

    # Con el tipo indicado resuelve al plan correcto.
    rows = [['maria@gym.cl', 'Pack', 'Plan 8 clases', '2026-06-01', '', 5]]
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_HEADERS)
    assert resp.status_code == 201, resp.content
    membership = StudentPlan.objects.get(user=membership_setup['student'])
    assert membership.plan.plan_type == 'pack'


def test_memberships_user_must_be_active_student_of_org(api_client, admin_a, org_b, make_user,
                                                        membership_setup):
    make_user('ajena', organization=org_b, role='student', email='ajena@gym.cl')
    make_user('coach2', organization=membership_setup['student'].organization,
              role='teacher', email='coach2@gym.cl')
    login(api_client, 'admin_a')
    rows = [
        ['ajena@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5],   # alumna de otra org
        ['coach2@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5],  # profesor, no alumno
    ]
    upload = build_xlsx(rows, headers=MEMBERSHIP_HEADERS)
    resp = api_client.post('/api/imports/memberships/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['summary']['errors'] == 2
    for row in body['rows']:
        assert row['errors'][0]['column'] == 'Email del alumno'
        assert 'No se encontró' in row['errors'][0]['message']


def test_memberships_inactive_membership_does_not_block(api_client, admin_a, membership_setup):
    import datetime

    from core.models import StudentPlan

    # Membresía VENCIDA/inactiva previa: no debe bloquear la importación.
    StudentPlan.objects.create(
        user=membership_setup['student'], plan=membership_setup['plan'],
        start_date=datetime.date(2025, 1, 1), end_date=datetime.date(2025, 1, 31),
        total_classes=8, classes_used=8, is_active=False,
    )
    login(api_client, 'admin_a')
    rows = [['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5]]
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_HEADERS)
    assert resp.status_code == 201
    assert resp.json()['created'] == 1
    assert StudentPlan.objects.filter(user=membership_setup['student']).count() == 2


def test_memberships_plan_of_other_org_not_found(api_client, admin_a, org_b, membership_setup):
    from core.models import Plan

    Plan.objects.create(organization=org_b, name='Plan ajeno', plan_type='monthly',
                        total_classes=8, duration_days=30, price=45000)
    login(api_client, 'admin_a')
    upload = build_xlsx([['maria@gym.cl', '', 'Plan ajeno', '2026-06-01', '', 5]],
                        headers=MEMBERSHIP_HEADERS)
    resp = api_client.post('/api/imports/memberships/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    error = body['rows'][0]['errors'][0]
    assert error['column'] == 'Nombre del plan'
    assert 'No se encontró' in error['message']


def test_memberships_inactive_student_not_found(api_client, admin_a, membership_setup):
    student = membership_setup['student']
    student.is_active = False
    student.save(update_fields=['is_active'])
    login(api_client, 'admin_a')
    upload = build_xlsx([['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5]],
                        headers=MEMBERSHIP_HEADERS)
    resp = api_client.post('/api/imports/memberships/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    assert 'No se encontró' in body['rows'][0]['errors'][0]['message']


def test_memberships_negative_remaining_rejected(api_client, admin_a, membership_setup):
    login(api_client, 'admin_a')
    upload = build_xlsx([['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', -1]],
                        headers=MEMBERSHIP_HEADERS)
    resp = api_client.post('/api/imports/memberships/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    assert 'entre 0 y 8' in body['rows'][0]['errors'][0]['message']


def test_memberships_preview_shows_derived_end_date(api_client, admin_a, membership_setup):
    login(api_client, 'admin_a')
    upload = build_xlsx([['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5]],
                        headers=MEMBERSHIP_HEADERS)
    resp = api_client.post('/api/imports/memberships/validate/', {'file': upload}, format='multipart')
    body = resp.json()
    # Fecha de término vacía: el preview muestra la fecha calculada (inicio + 29).
    assert body['rows'][0]['values']['Fecha de término'] == '2026-06-30'


def test_memberships_duplicate_student_in_file_first_wins(api_client, admin_a, membership_setup):
    from core.models import StudentPlan

    login(api_client, 'admin_a')
    rows = [
        ['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5],
        [' MARIA@GYM.CL ', '', 'Plan libre', '2026-06-01', '', ''],
    ]
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_HEADERS)
    assert resp.status_code == 201
    assert resp.json()['created'] == 1
    membership = StudentPlan.objects.get(user=membership_setup['student'])
    assert membership.plan_id == membership_setup['plan'].id  # la primera fila gana


# ----------------------------------------- Saldo flexible (utilizadas/restantes) + matrícula

MEMBERSHIP_FLEX_HEADERS = MEMBERSHIP_HEADERS + ['Clases utilizadas', 'Matrícula']


def _flex_validate(api_client, rows):
    upload = build_xlsx(rows, headers=MEMBERSHIP_FLEX_HEADERS)
    return api_client.post('/api/imports/memberships/validate/', {'file': upload}, format='multipart')


def test_memberships_import_used_only_computes_remaining(api_client, admin_a, membership_setup):
    from core.models import StudentPlan

    login(api_client, 'admin_a')
    # Solo 'Clases utilizadas' = 3 (restantes vacío) en un plan de 8.
    rows = [['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', '', 3, '']]
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_FLEX_HEADERS)
    assert resp.status_code == 201, resp.content
    membership = StudentPlan.objects.get(user=membership_setup['student'])
    assert membership.classes_used == 3
    assert membership.total_classes == 8


def test_memberships_import_both_consistent_ok(api_client, admin_a, membership_setup):
    from core.models import StudentPlan

    login(api_client, 'admin_a')
    rows = [['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5, 3, '']]  # 3 + 5 = 8
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_FLEX_HEADERS)
    assert resp.status_code == 201, resp.content
    membership = StudentPlan.objects.get(user=membership_setup['student'])
    assert membership.classes_used == 3


def test_memberships_import_both_inconsistent_rejected(api_client, admin_a, membership_setup):
    login(api_client, 'admin_a')
    rows = [['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5, 4, '']]  # 4 + 5 = 9 != 8
    body = _flex_validate(api_client, rows).json()
    assert body['can_commit'] is False
    assert 'No cuadra' in body['rows'][0]['errors'][0]['message']


def test_memberships_import_neither_balance_rejected(api_client, admin_a, membership_setup):
    login(api_client, 'admin_a')
    rows = [['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', '', '', '']]
    body = _flex_validate(api_client, rows).json()
    assert body['can_commit'] is False
    message = body['rows'][0]['errors'][0]['message']
    assert "'Clases restantes' o las 'Clases utilizadas'" in message


def test_memberships_import_used_out_of_range_rejected(api_client, admin_a, membership_setup):
    login(api_client, 'admin_a')
    rows = [['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', '', 10, '']]  # > 8
    body = _flex_validate(api_client, rows).json()
    assert body['can_commit'] is False
    assert 'entre 0 y 8' in body['rows'][0]['errors'][0]['message']


def test_memberships_import_sets_enrollment_fee(api_client, admin_a, membership_setup):
    from decimal import Decimal

    from core.models import StudentPlan

    login(api_client, 'admin_a')
    rows = [['maria@gym.cl', '', 'Plan 8 clases', '2026-06-01', '', 5, '', 50000]]
    resp = import_entity(api_client, 'memberships', rows, MEMBERSHIP_FLEX_HEADERS)
    assert resp.status_code == 201, resp.content
    membership = StudentPlan.objects.get(user=membership_setup['student'])
    assert membership.enrollment_fee == Decimal('50000')
    assert membership.enrollment_fee_due_at is not None  # vencimiento autocalculado


# ---------------------------------------------------------------- F5: Horario recurrente

TEMPLATE_HEADERS = ['Sucursal', 'Día de la semana', 'Hora de inicio', 'Hora de término',
                    'Nombre de la clase', 'Email del profesor', 'Tipo de clase', 'Disciplina',
                    'Capacidad', 'Vigente desde', 'Vigente hasta', 'Apta para clase de prueba']


@pytest.fixture
def schedule_setup(make_user, org_a):
    branch = Branch.objects.create(organization=org_a, name='Sede Centro')
    teacher = make_user('coach', organization=org_a, role='teacher', email='coach@gym.cl')
    class_type = ClassType.objects.create(organization=org_a, name='Clase grupal')
    discipline = Discipline.objects.create(organization=org_a, name='Yoga')
    return {'branch': branch, 'teacher': teacher, 'class_type': class_type,
            'discipline': discipline}


def test_class_templates_full_cycle(api_client, admin_a, org_a, schedule_setup):
    import datetime

    from core.models import ClassTemplate

    login(api_client, 'admin_a')
    rows = [
        ['Sede Centro', 'Lunes', '18:30', '19:30', 'Yoga vespertino', 'coach@gym.cl',
         'Clase grupal', 'Yoga', 15, '2026-06-15', '', 'Sí'],
        ['Sede Centro', 'Miércoles', '09:00', '10:00', '', '', '', '', '', '2026-06-15', '', ''],
    ]
    resp = import_entity(api_client, 'class-templates', rows, TEMPLATE_HEADERS)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 2

    yoga = ClassTemplate.objects.get(organization=org_a, name='Yoga vespertino')
    assert yoga.branch_id == schedule_setup['branch'].id
    assert yoga.weekday == 0
    assert yoga.start_time == datetime.time(18, 30)
    assert yoga.end_time == datetime.time(19, 30)
    assert yoga.teacher_id == schedule_setup['teacher'].id
    assert yoga.class_type_id == schedule_setup['class_type'].id
    assert yoga.discipline_id == schedule_setup['discipline'].id
    assert yoga.capacity == 15
    assert yoga.is_trial_eligible is True
    assert yoga.is_active is True

    plain = ClassTemplate.objects.get(organization=org_a, weekday=2)
    assert plain.teacher is None
    assert plain.capacity == 20  # default del spec
    assert plain.is_trial_eligible is False

    # Idempotente por (sucursal, día, hora de inicio, profesor).
    resp = import_entity(api_client, 'class-templates', rows, TEMPLATE_HEADERS)
    assert resp.json()['created'] == 0
    assert resp.json()['skipped_duplicates'] == 2


def test_class_templates_same_slot_different_teacher_both_import(
    api_client, admin_a, org_a, make_user, schedule_setup,
):
    import datetime

    from core.models import ClassTemplate

    coach2 = make_user('coach2', organization=org_a, role='teacher', email='coach2@gym.cl')
    login(api_client, 'admin_a')
    # Mismo branch / día / hora de inicio, profesores DISTINTOS → no se deduplican
    # (la clave natural incluye al profesor) y profesores distintos no se cruzan
    # entre sí → ambas se importan.
    rows = [
        ['Sede Centro', 'Lunes', '07:00', '08:00', 'Kick Boxing', 'coach@gym.cl',
         '', '', '', '2026-06-15', '', ''],
        ['Sede Centro', 'Lunes', '07:00', '08:00', 'Boxeo', 'coach2@gym.cl',
         '', '', '', '2026-06-15', '', ''],
    ]
    resp = import_entity(api_client, 'class-templates', rows, TEMPLATE_HEADERS)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 2
    assert resp.json()['skipped_duplicates'] == 0
    slot = dict(organization=org_a, branch=schedule_setup['branch'],
                weekday=0, start_time=datetime.time(7, 0))
    assert ClassTemplate.objects.filter(teacher=schedule_setup['teacher'], **slot).count() == 1
    assert ClassTemplate.objects.filter(teacher=coach2, **slot).count() == 1


def test_class_templates_same_slot_same_teacher_dedups(api_client, admin_a, org_a, schedule_setup):
    import datetime

    from core.models import ClassTemplate

    login(api_client, 'admin_a')
    # Mismo branch / día / hora de inicio y MISMO profesor → la segunda es duplicada
    # dentro del archivo y se omite.
    rows = [
        ['Sede Centro', 'Lunes', '07:00', '08:00', 'Kick Boxing', 'coach@gym.cl',
         '', '', '', '2026-06-15', '', ''],
        ['Sede Centro', 'Lunes', '07:00', '08:00', 'Repetida', 'coach@gym.cl',
         '', '', '', '2026-06-15', '', ''],
    ]
    resp = import_entity(api_client, 'class-templates', rows, TEMPLATE_HEADERS)
    assert resp.status_code == 201, resp.content
    assert resp.json()['created'] == 1
    assert resp.json()['skipped_duplicates'] == 1
    assert ClassTemplate.objects.filter(
        organization=org_a, weekday=0, start_time=datetime.time(7, 0),
    ).count() == 1


def test_class_templates_row_rules(api_client, admin_a, schedule_setup):
    login(api_client, 'admin_a')
    rows = [
        ['Sede Centro', 'Lunes', '19:30', '18:30', '', '', '', '', '', '2026-06-15', '', ''],
        ['Sede Centro', 'Martes', '18:30', '19:30', '', '', '', '', 0, '2026-06-15', '', ''],
        ['Sede Centro', 'Jueves', '18:30', '19:30', '', '', '', '', '', '2026-06-15', '2026-06-01', ''],
        ['Sede Centro', 'Lunsex', '18:30', '19:30', '', '', '', '', '', '2026-06-15', '', ''],
    ]
    upload = build_xlsx(rows, headers=TEMPLATE_HEADERS)
    resp = api_client.post('/api/imports/class-templates/validate/', {'file': upload},
                           format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    by_row = {r['row']: r for r in body['rows']}
    assert 'posterior a la hora de inicio' in by_row[2]['errors'][0]['message']
    assert 'entre 1 y 1000' in by_row[3]['errors'][0]['message']
    assert 'no puede ser anterior' in by_row[4]['errors'][0]['message']
    assert 'Las opciones son' in by_row[5]['errors'][0]['message']


def test_class_templates_teacher_overlap_against_db(api_client, admin_a, org_a, schedule_setup):
    import datetime

    from core.models import ClassTemplate

    ClassTemplate.objects.create(
        organization=org_a, branch=schedule_setup['branch'], teacher=schedule_setup['teacher'],
        weekday=0, start_time=datetime.time(18, 0), end_time=datetime.time(19, 0),
        start_date=datetime.date(2026, 6, 1), capacity=20,
    )
    login(api_client, 'admin_a')
    # 18:30-19:30 del lunes se cruza con la existente de 18:00-19:00.
    upload = build_xlsx(
        [['Sede Centro', 'Lunes', '18:30', '19:30', '', 'coach@gym.cl', '', '', '', '2026-06-15', '', '']],
        headers=TEMPLATE_HEADERS,
    )
    resp = api_client.post('/api/imports/class-templates/validate/', {'file': upload},
                           format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    error = body['rows'][0]['errors'][0]
    assert error['column'] == 'Email del profesor'
    assert 'se cruza con ese horario' in error['message']


def test_class_templates_teacher_overlap_within_file_rolls_back(api_client, admin_a, org_a,
                                                                schedule_setup):
    from core.models import ClassTemplate

    login(api_client, 'admin_a')
    # Dos filas del MISMO archivo se cruzan para el mismo profesor: el validate
    # no lo ve (no hay nada en BD), pero el commit lo atrapa con full_clean y
    # hace rollback total.
    rows = [
        ['Sede Centro', 'Lunes', '18:00', '19:00', '', 'coach@gym.cl', '', '', '', '2026-06-15', '', ''],
        ['Sede Centro', 'Lunes', '18:30', '19:30', '', 'coach@gym.cl', '', '', '', '2026-06-15', '', ''],
    ]
    file_bytes = build_xlsx_bytes(rows, headers=TEMPLATE_HEADERS)
    resp = api_client.post('/api/imports/class-templates/validate/',
                           {'file': as_upload(file_bytes)}, format='multipart')
    body = resp.json()
    assert body['can_commit'] is True  # el preview no puede verlo aún

    resp = api_client.post('/api/imports/class-templates/commit/',
                           {'file': as_upload(file_bytes), 'token': body['token']},
                           format='multipart')
    assert resp.status_code == 400
    assert 'No se importó ningún dato' in resp.json()['detail']
    error = resp.json()['rows'][0]['errors'][0]
    assert error['column'] == 'Email del profesor'
    assert ClassTemplate.objects.count() == 0  # rollback total


def test_class_templates_foreign_org_fks_not_found(api_client, admin_a, org_b, make_user,
                                                   schedule_setup):
    Branch.objects.create(organization=org_b, name='Sede Ajena')
    make_user('coach_b', organization=org_b, role='teacher', email='coach_b@gym.cl')
    login(api_client, 'admin_a')
    rows = [
        ['Sede Ajena', 'Lunes', '18:30', '19:30', '', '', '', '', '', '2026-06-15', '', ''],
        ['Sede Centro', 'Lunes', '18:30', '19:30', '', 'coach_b@gym.cl', '', '', '', '2026-06-15', '', ''],
    ]
    upload = build_xlsx(rows, headers=TEMPLATE_HEADERS)
    resp = api_client.post('/api/imports/class-templates/validate/', {'file': upload},
                           format='multipart')
    body = resp.json()
    assert body['summary']['errors'] == 2
    by_row = {r['row']: r for r in body['rows']}
    assert by_row[2]['errors'][0]['column'] == 'Sucursal'
    assert by_row[3]['errors'][0]['column'] == 'Email del profesor'
    for row in body['rows']:
        assert 'No se encontró' in row['errors'][0]['message']


def test_class_templates_commit_generates_calendar(api_client, admin_a, org_a, schedule_setup):
    from core.models import ClassTemplate, GymClass

    login(api_client, 'admin_a')
    # Vigencia acotada en el futuro lejano: 2030-01-07 y 2030-01-14 son lunes.
    rows = [['Sede Centro', 'Lunes', '18:30', '19:30', 'Yoga', 'coach@gym.cl',
             '', '', 10, '2030-01-07', '2030-01-20', '']]
    resp = import_entity(api_client, 'class-templates', rows, TEMPLATE_HEADERS)
    assert resp.status_code == 201, resp.content

    template = ClassTemplate.objects.get(organization=org_a, name='Yoga')
    assert template.created_by_id == admin_a.id  # trazabilidad del actor
    instances = GymClass.objects.filter(class_template=template).order_by('start_datetime')
    assert instances.count() == 2  # los dos lunes de la vigencia
    assert all(i.organization_id == org_a.id for i in instances)
    assert all(i.created_by_id == admin_a.id for i in instances)


def test_class_templates_teacher_overlap_against_calendar(api_client, admin_a, org_a,
                                                          schedule_setup):
    import datetime

    from django.utils import timezone as dj_tz

    from core.models import GymClass

    # Clase suelta (sin plantilla) del coach el lunes 2030-01-07 de 18:00 a 19:00.
    GymClass.objects.create(
        organization=org_a, branch=schedule_setup['branch'], teacher=schedule_setup['teacher'],
        name='Clase suelta',
        start_datetime=dj_tz.make_aware(datetime.datetime(2030, 1, 7, 18, 0)),
        end_datetime=dj_tz.make_aware(datetime.datetime(2030, 1, 7, 19, 0)),
    )
    login(api_client, 'admin_a')
    upload = build_xlsx(
        [['Sede Centro', 'Lunes', '18:30', '19:30', '', 'coach@gym.cl', '', '', '', '2030-01-01', '', '']],
        headers=TEMPLATE_HEADERS,
    )
    resp = api_client.post('/api/imports/class-templates/validate/', {'file': upload},
                           format='multipart')
    body = resp.json()
    assert body['can_commit'] is False
    error = body['rows'][0]['errors'][0]
    assert error['column'] == 'Email del profesor'
    assert 'clases en el calendario' in error['message']


def test_class_templates_template_references_include_weekdays(api_client, admin_a, schedule_setup):
    from openpyxl import load_workbook

    login(api_client, 'admin_a')
    resp = api_client.get('/api/imports/class-templates/template/')
    assert resp.status_code == 200
    references = load_workbook(BytesIO(resp.content))['Referencias']
    all_values = {
        references.cell(row=row, column=col).value
        for row in range(1, 12) for col in range(1, 10)
    }
    for label in ('Lunes', 'Miércoles', 'Domingo', 'Sede Centro', 'coach@gym.cl',
                  'Clase grupal', 'Yoga'):
        assert label in all_values


# ---------------------------------------------------------------- motor FK (unitario)

def test_resolve_fk_scoped_to_organization(org_a, org_b):
    from core.importer.fk import FKResolutionError, resolve_fk, reference_values
    from core.importer.spec import FKSpec

    fk = FKSpec(model='core.ClassType', lookup_field='name', reference_label='tipo de clase')
    funcional = ClassType.objects.create(organization=org_a, name='Funcional')
    ClassType.objects.create(organization=org_b, name='Funcional')  # otra org: invisible

    assert resolve_fk(fk, '  funcional ', org_a) == funcional

    with pytest.raises(FKResolutionError) as exc:
        resolve_fk(fk, 'Inexistente', org_a)
    assert "No se encontró tipo de clase 'Inexistente'" in exc.value.message

    # Ambigüedad dentro de la misma org (nombres que solo difieren en mayúsculas)
    ClassType.objects.create(organization=org_a, name='FUNCIONAL')
    with pytest.raises(FKResolutionError) as exc:
        resolve_fk(fk, 'funcional', org_a)
    assert 'más de un tipo de clase' in exc.value.message

    assert reference_values(fk, org_b) == ['Funcional']


# ---------------------------------------------------------------- F6: Upsert (motor)

def test_spec_upsert_contract():
    from core.importer.spec import (
        EntityImportSpec, FieldSpec, RowResult, ImportReport,
        STATUS_UPDATED, STATUS_UNCHANGED, STATUS_OK,
    )
    assert STATUS_UPDATED == 'actualizado'
    assert STATUS_UNCHANGED == 'sin_cambios'

    # FieldSpec.updatable default False, opt-in True
    assert FieldSpec(attr='a', label='A').updatable is False
    assert FieldSpec(attr='a', label='A', updatable=True).updatable is True

    # is_upsert: True si hay algún campo updatable o updatable_fields
    skip_spec = EntityImportSpec(slug='s', label='S', description='', model='core.Discipline',
                                 fields=(FieldSpec(attr='name', label='Nombre'),), natural_key=('name',))
    assert skip_spec.is_upsert is False
    up_spec = EntityImportSpec(slug='u', label='U', description='', model='core.Discipline',
                               fields=(FieldSpec(attr='name', label='Nombre', updatable=True),),
                               natural_key=('name',))
    assert up_spec.is_upsert is True
    up_spec2 = EntityImportSpec(slug='u2', label='U2', description='', model='core.Discipline',
                                fields=(FieldSpec(attr='name', label='Nombre'),),
                                natural_key=('name',), updatable_fields=('x',))
    assert up_spec2.is_upsert is True

    # RowResult.diff y conteos en el report
    report = ImportReport(rows=[
        RowResult(row=2, status=STATUS_OK, values={}),
        RowResult(row=3, status=STATUS_UPDATED, values={}, diff={'Clases utilizadas': {'from': 3, 'to': 5}}),
        RowResult(row=4, status=STATUS_UNCHANGED, values={}),
    ])
    assert report.updated == 1
    assert report.unchanged == 1
    assert report.summary()['updated'] == 1
    assert report.summary()['unchanged'] == 1
    assert report.can_commit is True  # updated/unchanged no son errores
    assert report.rows[1].diff == {'Clases utilizadas': {'from': 3, 'to': 5}}


def test_engine_existing_keys_returns_pk_map(org_a):
    from core.importer.engine import _existing_keys
    from core.importer.spec import EntityImportSpec, FieldSpec
    d = Discipline.objects.create(organization=org_a, name='Yoga')
    spec = EntityImportSpec(slug='d', label='D', description='', model='core.Discipline',
                            fields=(FieldSpec(attr='name', label='Nombre'),), natural_key=('name',))
    keys = _existing_keys(spec, org_a)
    assert keys == {('yoga',): d.pk}  # casefold + pk


def test_engine_diff_only_whitelist_and_changed(org_a):
    from core.importer.engine import _diff_updatable
    from core.importer.spec import EntityImportSpec, FieldSpec
    spec = EntityImportSpec(
        slug='d2', label='D2', description='', model='core.Discipline',
        fields=(FieldSpec(attr='name', label='Nombre'),
                FieldSpec(attr='description', label='Descripción', updatable=True)),
        natural_key=('name',),
    )
    existing = Discipline(name='Yoga', description='vieja')
    candidate = Discipline(name='Yoga-IGNORADO', description='nueva')  # name NO es updatable → no aparece
    diff = _diff_updatable(spec, existing, candidate)
    assert diff == {'Descripción': {'from': 'vieja', 'to': 'nueva'}}


def test_engine_validate_marks_updated_unchanged_and_first_wins(org_a):
    from core.importer import engine
    from core.importer.spec import (EntityImportSpec, FieldSpec,
                                     STATUS_UPDATED, STATUS_UNCHANGED, STATUS_DUP_FILE)
    spec = EntityImportSpec(
        slug='d3', label='D3', description='', model='core.Discipline',
        fields=(FieldSpec(attr='name', label='Nombre', required=True),
                FieldSpec(attr='description', label='Descripción', updatable=True)),
        natural_key=('name',),
    )
    Discipline.objects.create(organization=org_a, name='Yoga', description='vieja')
    Discipline.objects.create(organization=org_a, name='Pilates', description='igual')
    parsed = [
        (2, {'name': 'Yoga', 'description': 'nueva'}),     # UPDATED
        (3, {'name': 'Pilates', 'description': 'igual'}),  # UNCHANGED
        (4, {'name': 'Yoga', 'description': 'otra'}),      # repetida → DUP_FILE (primera gana)
    ]
    report = engine.validate_rows(spec, org_a, parsed)
    by_row = {r.row: r for r in report.rows}
    assert by_row[2].status == STATUS_UPDATED
    assert by_row[2].diff == {'Descripción': {'from': 'vieja', 'to': 'nueva'}}
    assert by_row[3].status == STATUS_UNCHANGED
    assert by_row[4].status == STATUS_DUP_FILE


def test_engine_commit_applies_only_whitelist(org_a):
    from core.importer import engine
    from core.importer.spec import EntityImportSpec, FieldSpec
    spec = EntityImportSpec(
        slug='d4', label='D4', description='', model='core.Discipline',
        fields=(FieldSpec(attr='name', label='Nombre', required=True),
                FieldSpec(attr='description', label='Descripción', updatable=True),
                FieldSpec(attr='is_active', label='Activa', kind='bool')),  # NO updatable
        natural_key=('name',),
    )
    d = Discipline.objects.create(organization=org_a, name='Yoga', description='vieja', is_active=True)
    rows = [['Yoga', 'nueva', 'No']]  # description cambia (updatable), is_active intenta cambiar (NO)
    file_bytes = build_xlsx_bytes(rows, headers=['Nombre', 'Descripción', 'Activa'])
    token = engine.issue_token(spec, org_a, file_bytes)
    report, created, updated = engine.run_commit(spec, org_a, as_upload(file_bytes), token)
    assert (created, updated) == (0, 1)
    d.refresh_from_db()
    assert d.description == 'nueva'   # whitelist aplicado
    assert d.is_active is True        # NO whitelist → intacto
