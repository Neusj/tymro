"""Generación de la plantilla XLSX de una entidad (hojas Datos / Instrucciones / Referencias).

La hoja Referencias se puebla con los valores reales de LA organización del
actor (FK) más los vocabularios fijos (Sí/No, choices), y la hoja Datos usa
dropdowns de Excel (DataValidation) apuntando a esos rangos.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .fk import reference_values

HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='F97316', end_color='F97316', fill_type='solid')
EXAMPLE_FONT = Font(italic=True, color='999999')
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True)
DROPDOWN_ROWS = 1000  # filas de datos con dropdown habilitado


def _set_widths(sheet, widths):
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _vocabularies(spec, organization):
    """[(field, título del bloque, valores)] para Referencias y dropdowns."""
    blocks = []
    for field in spec.fields:
        if field.kind == 'fk':
            values = reference_values(field.fk, organization)
            title = f'{field.label} (valores válidos)'
            blocks.append((field, title, [str(v) for v in values]))
        elif field.kind == 'choice':
            blocks.append((field, field.label, [str(label) for label, _ in field.choices]))
        elif field.kind == 'bool':
            blocks.append((field, field.label, ['Sí', 'No']))
    return blocks


def build_template(spec, organization):
    workbook = Workbook()

    # --- Hoja Datos ---
    data_sheet = workbook.active
    data_sheet.title = 'Datos'
    labels = [field.label for field in spec.fields]
    data_sheet.append(labels)
    for cell in data_sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    example_row = [field.example for field in spec.fields]
    if any(example_row):
        data_sheet.append(example_row)
        for cell in data_sheet[2]:
            cell.font = EXAMPLE_FONT
    _set_widths(data_sheet, [max(len(f.label) + 4, len(str(f.example)) + 4, 14) for f in spec.fields])
    data_sheet.freeze_panes = 'A2'

    # --- Hoja Referencias (antes que los dropdowns, que apuntan a sus rangos) ---
    vocab_blocks = _vocabularies(spec, organization)
    ref_sheet = workbook.create_sheet('Referencias')
    ref_sheet.append(['Valores válidos para las columnas con lista desplegable.'])
    ref_sheet['A1'].font = SECTION_FONT
    ref_ranges = {}
    widths = []
    for col_idx, (field, title, values) in enumerate(vocab_blocks, start=1):
        letter = get_column_letter(col_idx)
        ref_sheet.cell(row=2, column=col_idx, value=title).font = SECTION_FONT
        for row_offset, value in enumerate(values, start=3):
            cell = ref_sheet.cell(row=row_offset, column=col_idx, value=value)
            # Los valores vienen de datos de la org: forzar texto plano para que
            # un nombre tipo "=HYPERLINK(...)" nunca se interprete como fórmula.
            cell.data_type = 's'
        last_row = max(3, 2 + len(values))
        ref_ranges[field.attr] = f'Referencias!${letter}$3:${letter}${last_row}'
        widths.append(max([len(title) + 4] + [len(v) + 4 for v in values] + [14]))
    if not vocab_blocks:
        ref_sheet.cell(row=2, column=1, value='Esta entidad no tiene columnas con lista desplegable.')
        widths = [60]
    _set_widths(ref_sheet, widths)

    # Dropdowns en Datos apuntando a Referencias
    for col_idx, field in enumerate(spec.fields, start=1):
        formula = ref_ranges.get(field.attr)
        if not formula:
            continue
        validation = DataValidation(
            type='list', formula1=f'={formula}', allow_blank=True, showDropDown=False,
            error='Elige un valor de la lista (hoja Referencias).',
            errorTitle='Valor no válido',
        )
        data_sheet.add_data_validation(validation)
        letter = get_column_letter(col_idx)
        validation.add(f'{letter}2:{letter}{DROPDOWN_ROWS + 1}')

    # --- Hoja Instrucciones ---
    info_sheet = workbook.create_sheet('Instrucciones')
    info_sheet['A1'] = f'Cómo importar {spec.label}'
    info_sheet['A1'].font = TITLE_FONT
    row = 3
    general_steps = (
        'Completa la hoja "Datos": una fila por registro, sin tocar los encabezados.',
        'Las columnas con lista desplegable solo aceptan los valores de la hoja "Referencias".',
        'Cuando termines, guarda el archivo y súbelo en TYMRO: primero se valida y te '
        'mostraremos un resumen; nada se guarda hasta que confirmes.',
        'Si alguna fila tiene errores no se importará NADA: corrige y vuelve a subir.',
    )
    for text in general_steps + tuple(spec.instructions):
        info_sheet.cell(row=row, column=1, value=f'• {text}')
        row += 1

    row += 1
    info_sheet.cell(row=row, column=1, value='Detalle de las columnas').font = SECTION_FONT
    row += 1
    header = ('Columna', '¿Obligatoria?', 'Descripción', 'Ejemplo', 'Valores permitidos')
    for col_idx, text in enumerate(header, start=1):
        info_sheet.cell(row=row, column=col_idx, value=text).font = SECTION_FONT
    row += 1
    for field in spec.fields:
        if field.kind == 'fk':
            allowed = 'Elige un valor de la hoja "Referencias".'
        elif field.kind == 'choice':
            allowed = ', '.join(str(label) for label, _ in field.choices)
        elif field.kind == 'bool':
            allowed = 'Sí / No'
        elif field.kind == 'date':
            allowed = 'Fecha AAAA-MM-DD (ej. 2026-03-01)'
        elif field.kind == 'time':
            allowed = 'Hora HH:MM (ej. 18:30)'
        else:
            allowed = 'Texto libre' if field.kind in ('string', 'text') else ''
        if field.max_length:
            allowed = f'{allowed} (máx. {field.max_length} caracteres)'.strip()
        values = (
            field.label,
            'Sí' if field.required else 'No',
            field.help_text,
            str(field.example),
            allowed,
        )
        for col_idx, text in enumerate(values, start=1):
            info_sheet.cell(row=row, column=col_idx, value=text)
        row += 1
    _set_widths(info_sheet, [28, 14, 55, 24, 40])

    return workbook
