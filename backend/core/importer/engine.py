"""Motor genérico de importación: parseo XLSX, pipeline por fila, tokens y commit.

El motor no conoce modelos concretos: todo viene del ``EntityImportSpec``.
La organización destino llega SIEMPRE resuelta desde la view (regla #1 de
multitenancy); el archivo no puede aportarla ni cambiarla.

Flujo validate → commit: ``validate`` no persiste nada y emite un token firmado
(entidad + organización + sha256 del archivo). ``commit`` exige el mismo archivo
más el token, re-valida TODO de cero y persiste dentro de ``transaction.atomic()``:
cualquier error implica rollback total (sin datos fantasma).
"""
import datetime
import hashlib
from io import BytesIO

from django.core import signing
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import IntegrityError, transaction

from .fk import FKResolutionError, resolve_fk
from .spec import (
    STATUS_DUP_DB,
    STATUS_DUP_FILE,
    STATUS_ERROR,
    STATUS_OK,
    STATUS_UNCHANGED,
    STATUS_UPDATED,
    ImportReport,
    RowError,
    RowResult,
)

MAX_FILE_SIZE = 5 * 1024 * 1024
ALLOWED_EXTENSION = '.xlsx'
DATA_SHEET = 'Datos'
IMPORT_TOKEN_SALT = 'core.importer.preview'
IMPORT_TOKEN_MAX_AGE = 60 * 30  # 30 minutos

TRUE_WORDS = {'sí', 'si', '1', 'true', 'verdadero'}
FALSE_WORDS = {'no', '0', 'false', 'falso'}


class ImportFileError(Exception):
    """Problema a nivel de archivo (no de fila): la view responde 400."""

    def __init__(self, message):
        self.message = message
        super().__init__(message)


class ImportCommitError(Exception):
    """El commit re-validó y encontró errores: rollback total."""

    def __init__(self, report):
        self.report = report
        super().__init__('La importación tiene errores')


def _normalize_header(text):
    return str(text or '').strip().casefold()


def _to_text(value):
    """Texto mostrable/limpio de una celda (Excel entrega números como float)."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, bool):
        value = 'Sí' if value else 'No'
    return str(value).strip()


def read_upload(uploaded_file):
    """Valida tamaño y extensión; devuelve los bytes (sirven para hash y parseo)."""
    if uploaded_file is None:
        raise ImportFileError('No se recibió ningún archivo. Adjunta la plantilla completada.')
    name = (uploaded_file.name or '').lower()
    if not name.endswith(ALLOWED_EXTENSION):
        raise ImportFileError('El archivo debe ser un Excel (.xlsx). Descarga la plantilla y úsala como base.')
    if uploaded_file.size > MAX_FILE_SIZE:
        raise ImportFileError('El archivo supera el tamaño máximo de 5 MB.')
    return uploaded_file.read()


def parse_workbook(spec, file_bytes):
    """Lee la hoja "Datos" y devuelve [(nº fila Excel, {attr: valor crudo})]."""
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        raise ImportFileError(
            'El archivo no es un Excel válido (.xlsx). Descarga la plantilla y úsala como base.'
        ) from None

    sheet = None
    for ws in workbook.worksheets:
        if _normalize_header(ws.title) == _normalize_header(DATA_SHEET):
            sheet = ws
            break
    if sheet is None:
        raise ImportFileError(
            f'El archivo no tiene la hoja "{DATA_SHEET}". Usa la plantilla descargada como base.'
        )

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ImportFileError('La hoja "Datos" está vacía.') from None

    header_index = {}
    for idx, header in enumerate(header_row or ()):
        normalized = _normalize_header(header)
        if normalized and normalized not in header_index:
            header_index[normalized] = idx

    column_map = {}
    for field in spec.fields:
        normalized = _normalize_header(field.label)
        if normalized in header_index:
            column_map[field.attr] = header_index[normalized]
        elif field.required:
            raise ImportFileError(
                f"Falta la columna '{field.label}'. Usa la plantilla descargada "
                'sin modificar los encabezados.'
            )
        # Columna OPCIONAL ausente: se trata como celdas vacías (cada fila la
        # recibe como None). Así agregar un campo opcional no rompe las
        # plantillas ya distribuidas (importador flexible).

    parsed = []
    physical_rows = 0
    for row_number, row in enumerate(rows_iter, start=2):
        # Tope de filas FÍSICAS (incluidas las vacías): evita iterar un XLSX
        # artesanal con millones de filas declaradas que sí cabe en 5 MB.
        physical_rows += 1
        if physical_rows > spec.max_rows * 10:
            raise ImportFileError(
                'El archivo tiene demasiadas filas (incluso vacías). '
                'Usa la plantilla descargada y divide la carga en varios archivos.'
            )
        raw = {
            attr: (row[idx] if row is not None and idx < len(row) else None)
            for attr, idx in column_map.items()
        }
        if all(value is None or str(value).strip() == '' for value in raw.values()):
            continue
        parsed.append((row_number, raw))
        if len(parsed) > spec.max_rows:
            raise ImportFileError(
                f'El archivo tiene más de {spec.max_rows} filas con datos. '
                'Divide la carga en varios archivos.'
            )
    return parsed


def _coerce(field, raw):
    """Convierte el valor crudo de la celda al tipo del campo.

    Devuelve (valor, None) o (None, mensaje de error en español).
    Vacío => (None, None): requeridos/default se resuelven después.
    """
    if raw is None or str(raw).strip() == '':
        return None, None

    kind = field.kind
    if kind in ('string', 'text', 'fk'):
        return _to_text(raw), None

    if kind == 'email':
        text = _to_text(raw)
        try:
            validate_email(text)
        except DjangoValidationError:
            return None, f"El correo '{text}' de '{field.label}' no es válido. Revisa el formato (ej. nombre@dominio.cl)."
        return text, None

    if kind == 'bool':
        if isinstance(raw, bool):
            return raw, None
        text = _to_text(raw).casefold()
        if text in TRUE_WORDS:
            return True, None
        if text in FALSE_WORDS:
            return False, None
        return None, (
            f"El valor '{_to_text(raw)}' no es válido para '{field.label}'. Usa 'Sí' o 'No'."
        )

    if kind == 'int':
        if isinstance(raw, bool):
            raw = int(raw)
        try:
            value = float(raw)
        except (TypeError, ValueError):
            return None, f"El valor '{_to_text(raw)}' no es un número válido para '{field.label}'."
        if not value.is_integer():
            return None, f"El valor de '{field.label}' debe ser un número entero (sin decimales)."
        return int(value), None

    if kind == 'decimal':
        try:
            return float(raw), None
        except (TypeError, ValueError):
            return None, f"El valor '{_to_text(raw)}' no es un número válido para '{field.label}'."

    if kind == 'date':
        if isinstance(raw, datetime.datetime):
            return raw.date(), None
        if isinstance(raw, datetime.date):
            return raw, None
        try:
            return datetime.date.fromisoformat(_to_text(raw)), None
        except ValueError:
            return None, (
                f"La fecha de '{field.label}' no es válida. Usa el formato AAAA-MM-DD, "
                'por ejemplo 2026-03-01.'
            )

    if kind == 'time':
        if isinstance(raw, datetime.datetime):
            return raw.time(), None
        if isinstance(raw, datetime.time):
            return raw, None
        try:
            return datetime.time.fromisoformat(_to_text(raw)), None
        except ValueError:
            return None, (
                f"La hora de '{field.label}' no es válida. Usa el formato HH:MM, por ejemplo 18:30."
            )

    if kind == 'choice':
        text = _to_text(raw).casefold()
        for choice_label, db_value in field.choices:
            if str(choice_label).casefold() == text or str(db_value).casefold() == text:
                return db_value, None
        allowed = ', '.join(f"'{label}'" for label, _ in field.choices)
        return None, (
            f"El valor '{_to_text(raw)}' no es válido para '{field.label}'. "
            f'Las opciones son: {allowed}.'
        )

    return _to_text(raw), None


def _display_value(field, raw):
    if raw is None:
        return ''
    if field.kind == 'date' and isinstance(raw, datetime.date):
        # Preview en formato chileno dd-mm-yyyy (la entrada sigue siendo AAAA-MM-DD).
        return raw.strftime('%d-%m-%Y')
    if field.kind == 'bool' and isinstance(raw, bool):
        return 'Sí' if raw else 'No'
    if field.kind == 'choice':
        for choice_label, db_value in field.choices:
            if db_value == raw:
                return str(choice_label)
    if field.kind == 'fk' and field.fk and hasattr(raw, field.fk.lookup_field):
        # Mostrar el valor de búsqueda (nombre/email), no el __str__ del modelo.
        return str(getattr(raw, field.fk.lookup_field))
    return _to_text(raw)


def _natural_key_from_cleaned(spec, cleaned):
    parts = []
    for attr in spec.natural_key:
        value = cleaned.get(attr)
        if hasattr(value, 'pk'):
            parts.append(value.pk)
        elif isinstance(value, str):
            parts.append(value.strip().casefold())
        else:
            parts.append(value)
    return tuple(parts)


def _existing_keys(spec, organization):
    """{clave_natural: pk} ya existentes en la BD de ESTA organización.

    La clave excluye el pk; el pk se usa en upsert para traer el objeto a
    actualizar. Para specs skip, solo se consulta la pertenencia (``key in ...``).
    """
    from django.apps import apps

    model = apps.get_model(spec.model)
    columns = []
    for attr in spec.natural_key:
        field = spec.field_by_attr(attr)
        columns.append(f'{attr}_id' if field.kind == 'fk' else attr)
    existing = {}
    queryset = model.objects.filter(**{spec.org_field: organization})
    if spec.dedup_filters:
        queryset = queryset.filter(**spec.dedup_filters)
    for values in queryset.values_list(*columns, 'pk'):
        *key_values, pk = values
        key = tuple(
            value.strip().casefold() if isinstance(value, str) else value
            for value in key_values
        )
        existing[key] = pk
    return existing


def _build_candidate(spec, organization, kwargs):
    """Instancia (sin guardar) que representa lo que el archivo quiere persistir.

    Usa el ``build_instance`` del spec si existe; si no, construye el modelo
    fijando la organización del actor (mismo camino que el create del commit).
    En validate corre fuera de transacción: los specs con guardas que usen
    ``select_for_update`` deben gatearlas en ``connection.in_atomic_block``.
    """
    from django.apps import apps

    if spec.build_instance:
        return spec.build_instance(kwargs, organization)
    model = apps.get_model(spec.model)
    return model(**{spec.org_field: organization}, **kwargs)


def _updatable_attrs(spec):
    """Whitelist de atributos del modelo que el upsert compara y aplica."""
    return tuple(f.attr for f in spec.fields if f.updatable) + tuple(spec.updatable_fields)


def _diff_updatable(spec, existing, candidate):
    """{label_es: {'from','to'}} de atributos whitelisted que existen en el
    candidato y difieren del objeto en BD. Atributos no presentes en el
    candidato (ej. 'remaining_classes', que no es campo del modelo) se ignoran."""
    diff = {}
    for attr in _updatable_attrs(spec):
        if not hasattr(candidate, attr):
            continue
        new_value = getattr(candidate, attr)
        old_value = getattr(existing, attr, None)
        if old_value != new_value:
            try:
                label = spec.field_by_attr(attr).label
            except KeyError:
                label = attr  # derivado sin columna (ej. 'is_active')
            diff[label] = {'from': old_value, 'to': new_value}
    return diff


def validate_rows(spec, organization, parsed):
    """Pipeline por fila: coerción → requeridos → FK → dedup → reglas → derive."""
    report = ImportReport()
    existing_keys = _existing_keys(spec, organization)
    seen_keys = {}

    for row_number, raw in parsed:
        errors = []
        cleaned = {}
        display = {}
        pending_fks = []  # las FK se resuelven en una 2ª pasada, cuando los
        # posibles disambiguators (otras columnas) ya están coercionados.

        for field in spec.fields:
            raw_value = raw.get(field.attr)
            value, error = _coerce(field, raw_value)
            if error:
                errors.append(RowError(row=row_number, column=field.label, message=error))
                display[field.label] = _to_text(raw_value)
                continue

            if value is None:
                if field.required:
                    errors.append(RowError(
                        row=row_number, column=field.label,
                        message=f"El campo '{field.label}' es obligatorio.",
                    ))
                    display[field.label] = ''
                    continue
                value = field.default
                if value is None:
                    display[field.label] = ''
                    cleaned[field.attr] = '' if field.kind in ('string', 'text', 'email') else None
                    continue

            if field.max_length and isinstance(value, str) and len(value) > field.max_length:
                errors.append(RowError(
                    row=row_number, column=field.label,
                    message=f"El campo '{field.label}' supera el máximo de {field.max_length} caracteres.",
                ))
                display[field.label] = _to_text(raw_value)
                continue

            if field.kind == 'fk':
                pending_fks.append((field, value))
                display[field.label] = _to_text(raw_value)
                continue

            cleaned[field.attr] = value
            display[field.label] = _display_value(field, value)

        for field, text_value in pending_fks:
            extra_filters = {}
            for dis_attr, dis_lookup in (field.fk.disambiguators or ()):
                if cleaned.get(dis_attr) is not None:
                    extra_filters[dis_lookup] = cleaned[dis_attr]
            try:
                value = resolve_fk(field.fk, text_value, organization, extra_filters=extra_filters)
            except FKResolutionError as exc:
                errors.append(RowError(row=row_number, column=field.label, message=exc.message))
                continue
            cleaned[field.attr] = value
            display[field.label] = _display_value(field, value)

        if errors:
            report.rows.append(RowResult(
                row=row_number, status=STATUS_ERROR, values=display, errors=errors,
            ))
            continue

        # Dedup ANTES de las reglas de fila para el caso skip: una fila que se va
        # a omitir no se valida (si no, re-importar el mismo archivo marcaría como
        # conflicto contra BD lo que es su propio duplicado, rompiendo la
        # idempotencia). OJO: la clave natural no puede depender de valores de derive.
        # En upsert (spec.is_upsert) una fila existente NO se omite: se evalúa para
        # actualizar, así que sí pasa por reglas y derive.
        key = _natural_key_from_cleaned(spec, cleaned)
        key_labels = ' / '.join(display.get(spec.field_by_attr(a).label, '') for a in spec.natural_key)
        if key in seen_keys:  # primera fila gana (OK / UPDATED / UNCHANGED)
            report.rows.append(RowResult(
                row=row_number, status=STATUS_DUP_FILE, values=display,
                note=f'Repetida dentro del archivo (igual que la fila {seen_keys[key]}): se omitirá.',
            ))
            continue
        in_db = key in existing_keys
        if in_db and not spec.is_upsert:  # skip clásico: no se toca lo existente
            report.rows.append(RowResult(
                row=row_number, status=STATUS_DUP_DB, values=display,
                note=f"Ya existe '{key_labels}' en tu organización: se omitirá.",
            ))
            continue

        for validator in spec.row_validators:
            for error in validator(cleaned, organization) or []:
                # Se reconstruye el RowError para no depender de mutar el
                # objeto que devolvió el spec (el spec no conoce la fila).
                errors.append(RowError(row=row_number, column=error.column, message=error.message))
        if errors:
            report.rows.append(RowResult(
                row=row_number, status=STATUS_ERROR, values=display, errors=errors,
            ))
            continue

        # derive se aplica ANTES del preview para que el usuario confirme
        # exactamente lo que se va a guardar (ej. plan ilimitado → 0 clases).
        if spec.derive:
            for attr, value in (spec.derive(cleaned, organization) or {}).items():
                cleaned[attr] = value
                try:
                    derived_field = spec.field_by_attr(attr)
                except KeyError:
                    continue
                display[derived_field.label] = _display_value(derived_field, value)

        if in_db:  # spec.is_upsert garantizado: comparar contra el objeto existente
            from django.apps import apps

            existing = apps.get_model(spec.model).objects.get(pk=existing_keys[key])
            candidate = _build_candidate(spec, organization, dict(cleaned))
            diff = _diff_updatable(spec, existing, candidate)
            seen_keys[key] = row_number
            if diff:
                report.rows.append(RowResult(
                    row=row_number, status=STATUS_UPDATED, values=display,
                    cleaned=cleaned, diff=diff, note='Se actualizará.',
                ))
            else:
                report.rows.append(RowResult(
                    row=row_number, status=STATUS_UNCHANGED, values=display,
                    note='Sin cambios: se omitirá.',
                ))
            continue

        seen_keys[key] = row_number
        report.rows.append(RowResult(
            row=row_number, status=STATUS_OK, values=display, cleaned=cleaned,
        ))

    return report


def issue_token(spec, organization, file_bytes):
    payload = {
        'entity': spec.slug,
        'organization_id': organization.pk,
        'sha256': hashlib.sha256(file_bytes).hexdigest(),
    }
    return signing.dumps(payload, salt=IMPORT_TOKEN_SALT)


def verify_token(token, spec, organization, file_bytes):
    if not token:
        raise ImportFileError('Falta el token de previsualización. Primero valida el archivo.')
    try:
        payload = signing.loads(token, salt=IMPORT_TOKEN_SALT, max_age=IMPORT_TOKEN_MAX_AGE)
    except signing.SignatureExpired:
        raise ImportFileError(
            'La previsualización expiró (30 minutos). Vuelve a validar el archivo.'
        ) from None
    except signing.BadSignature:
        raise ImportFileError(
            'El token de previsualización no es válido. Vuelve a validar el archivo.'
        ) from None

    expected = {
        'entity': spec.slug,
        'organization_id': organization.pk,
        'sha256': hashlib.sha256(file_bytes).hexdigest(),
    }
    if payload != expected:
        raise ImportFileError(
            'El archivo no coincide con el que validaste. Vuelve a validar antes de confirmar.'
        )


def run_validate(spec, organization, uploaded_file):
    """Valida todo sin persistir. Devuelve (report, token)."""
    file_bytes = read_upload(uploaded_file)
    parsed = parse_workbook(spec, file_bytes)
    report = validate_rows(spec, organization, parsed)
    return report, issue_token(spec, organization, file_bytes)


def _instance_org_id(instance, org_field):
    """organization_id de la instancia siguiendo rutas con '__' (ej. user__organization)."""
    *path, last = org_field.split('__')
    obj = instance
    for part in path:
        obj = getattr(obj, part, None)
        if obj is None:
            return None
    return getattr(obj, f'{last}_id', None)


def _field_name_for_attr(model, attr):
    """Nombre de campo para save(update_fields=...) a partir de un attr.

    Maneja FKs: 'plan_id' (attname) → 'plan' (name). Devuelve None si el attr
    no corresponde a un campo concreto del modelo.
    """
    for f in model._meta.concrete_fields:
        if attr in (f.attname, f.name):
            return f.name
    return None


def _model_validation_errors(spec, row_number, exc):
    """Convierte un ValidationError de full_clean() en RowErrors con label español."""
    error_dict = getattr(exc, 'message_dict', None) or {'__all__': exc.messages}
    errors = []
    for attr, messages in error_dict.items():
        try:
            column = spec.field_by_attr(attr).label
        except KeyError:
            column = ''
        for message in messages:
            errors.append(RowError(row=row_number, column=column, message=str(message)))
    return errors


def run_commit(spec, organization, uploaded_file, token, actor=None):
    """Re-valida de cero y persiste atómicamente. Devuelve (report, created, updated)."""
    from django.apps import apps

    file_bytes = read_upload(uploaded_file)
    verify_token(token, spec, organization, file_bytes)
    parsed = parse_workbook(spec, file_bytes)

    model = apps.get_model(spec.model)
    try:
        with transaction.atomic():
            report = validate_rows(spec, organization, parsed)
            if not report.can_commit:
                raise ImportCommitError(report)
            # Mapa clave→pk para las filas a actualizar (mismo dict que usó validate).
            existing_keys_commit = _existing_keys(spec, organization) if spec.is_upsert else {}
            created = 0
            updated = 0
            created_instances = []
            for row in report.rows:
                if row.status not in (STATUS_OK, STATUS_UPDATED):
                    continue
                # derive ya fue aplicado por validate_rows sobre row.cleaned.
                kwargs = dict(row.cleaned)
                if spec.org_field in kwargs:
                    raise RuntimeError(
                        f'El spec "{spec.slug}" intentó fijar {spec.org_field} (multitenancy regla #1)'
                    )

                if row.status == STATUS_UPDATED:
                    key = _natural_key_from_cleaned(spec, row.cleaned)
                    existing = model.objects.select_for_update().get(pk=existing_keys_commit[key])
                    # Red de seguridad multitenant: jamás actualizar fuera de la org del actor.
                    if _instance_org_id(existing, spec.org_field) != organization.pk:
                        raise RuntimeError(f'update de "{spec.slug}" cruzaría organización')
                    # Candidato en atomic: el spec corre aquí su guarda/lock (in_atomic_block).
                    candidate = _build_candidate(spec, organization, kwargs)
                    update_fields = []
                    for attr in _updatable_attrs(spec):
                        if not hasattr(candidate, attr):
                            continue
                        setattr(existing, attr, getattr(candidate, attr))
                        field_name = _field_name_for_attr(model, attr)
                        if field_name:
                            update_fields.append(field_name)
                    try:
                        existing.full_clean()
                    except DjangoValidationError as exc:
                        row.errors = _model_validation_errors(spec, row.row, exc)
                        row.status = STATUS_ERROR
                        raise ImportCommitError(report) from None
                    if hasattr(existing, 'updated_at'):
                        update_fields.append('updated_at')
                    existing.save(update_fields=update_fields)
                    updated += 1
                    continue

                if spec.build_instance:
                    instance = spec.build_instance(kwargs, organization)
                    # Red de seguridad multitenant: un spec jamás puede crear
                    # instancias fuera de la organización del actor.
                    if _instance_org_id(instance, spec.org_field) != organization.pk:
                        raise RuntimeError(
                            f'build_instance de "{spec.slug}" no asignó la organización del actor'
                        )
                else:
                    instance = model(**{spec.org_field: organization}, **kwargs)
                # Trazabilidad: si el modelo registra autor, dejar al actor real.
                if actor is not None and hasattr(instance, 'created_by_id') and instance.created_by_id is None:
                    instance.created_by = actor
                try:
                    instance.full_clean()
                except DjangoValidationError as exc:
                    row.errors = _model_validation_errors(spec, row.row, exc)
                    row.status = STATUS_ERROR
                    raise ImportCommitError(report) from None
                instance.save()
                created_instances.append(instance)
                created += 1
            # post_commit SOLO para creadas (las actualizaciones no lo disparan).
            if spec.post_commit and created_instances:
                spec.post_commit(created_instances, organization, actor)
    except IntegrityError:
        raise ImportFileError(
            'Otro proceso modificó los datos mientras importabas. '
            'Vuelve a validar el archivo e inténtalo de nuevo.'
        ) from None
    return report, created, updated
